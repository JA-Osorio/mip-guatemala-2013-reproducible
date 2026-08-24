from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import MipSourceData, SourceLayout


FINAL_COMPONENT_COLUMNS = {
    "exportaciones_fob": "exports_column",
    "consumo_hogares": "households_column",
    "consumo_isflsh": "npish_column",
    "consumo_gobierno": "government_column",
    "formacion_bruta_capital_fijo": "gfcf_column",
    "variacion_existencias": "inventories_column",
    "ajuste_cif_fob": "cif_fob_column",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        raise TypeError("Se encontró un booleano donde se esperaba un valor numérico.")
    return float(value)


def _vector_by_rows(ws: Worksheet, rows: Iterable[int], column: int) -> np.ndarray:
    return np.asarray([_number(ws.cell(row, column).value) for row in rows], dtype=float)


def _vector_by_columns(ws: Worksheet, row: int, columns: Iterable[int]) -> np.ndarray:
    return np.asarray([_number(ws.cell(row, column).value) for column in columns], dtype=float)


def _matrix(ws: Worksheet, rows: Iterable[int], columns: Iterable[int]) -> np.ndarray:
    column_list = list(columns)
    return np.asarray(
        [[_number(ws.cell(row, column).value) for column in column_list] for row in rows],
        dtype=float,
    )


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _final_components(
    ws: Worksheet,
    rows: range,
    layout: SourceLayout,
) -> dict[str, np.ndarray]:
    return {
        name: _vector_by_rows(ws, rows, getattr(layout, layout_field))
        for name, layout_field in FINAL_COMPONENT_COLUMNS.items()
    }


def extract_source(path: Path, layout: SourceLayout) -> MipSourceData:
    if not path.exists():
        raise FileNotFoundError(
            f"No se encontró la fuente {path}. Consulte "
            "00_trazabilidad_fuentes/instrucciones_fuente_original.txt."
        )

    workbook = load_workbook(path, data_only=True, read_only=False, keep_links=False)
    try:
        ws_domestic = workbook[layout.sheet_domestic]
        ws_imported = workbook[layout.sheet_imported]

        rows = range(layout.product_first_row, layout.product_last_row + 1)
        columns = range(layout.matrix_first_column, layout.matrix_last_column + 1)

        codes = tuple(_text(ws_domestic.cell(row, layout.code_column).value) for row in rows)
        labels = tuple(_text(ws_domestic.cell(row, layout.label_column).value) for row in rows)
        column_codes = tuple(
            _text(ws_domestic.cell(layout.product_header_row, column).value)
            for column in columns
        )
        imported_codes = tuple(
            _text(ws_imported.cell(row, layout.code_column).value) for row in rows
        )
        imported_column_codes = tuple(
            _text(ws_imported.cell(layout.product_header_row, column).value)
            for column in columns
        )
        if imported_codes != codes or imported_column_codes != column_codes:
            raise ValueError("Las nomenclaturas de las dos hojas de la fuente no coinciden.")

        return MipSourceData(
            source_path=path,
            source_sha256=sha256_file(path),
            codes=codes,
            labels=labels,
            column_codes=column_codes,
            z_domestic=_matrix(ws_domestic, rows, columns),
            z_imported=_matrix(ws_imported, rows, columns),
            final_domestic=_final_components(ws_domestic, rows, layout),
            final_imported=_final_components(ws_imported, rows, layout),
            total_intermediate_domestic_source=_vector_by_rows(
                ws_domestic, rows, layout.total_intermediate_column
            ),
            total_intermediate_imported_source=_vector_by_rows(
                ws_imported, rows, layout.total_intermediate_column
            ),
            total_utilization_domestic=_vector_by_rows(
                ws_domestic, rows, layout.total_utilization_column
            ),
            total_utilization_imported=_vector_by_rows(
                ws_imported, rows, layout.total_utilization_column
            ),
            domestic_intermediate_by_column_source=_vector_by_columns(
                ws_domestic, layout.row_domestic_intermediate, columns
            ),
            imported_intermediate_by_column_source=_vector_by_columns(
                ws_domestic, layout.row_imported_intermediate, columns
            ),
            taxes_products=_vector_by_columns(ws_domestic, layout.row_taxes_products, columns),
            subsidies_products=_vector_by_columns(
                ws_domestic, layout.row_subsidies_products, columns
            ),
            net_taxes_products=_vector_by_columns(
                ws_domestic, layout.row_net_taxes_products, columns
            ),
            value_added=_vector_by_columns(ws_domestic, layout.row_value_added, columns),
            output=_vector_by_columns(ws_domestic, layout.row_output, columns),
            gdp=_vector_by_columns(ws_domestic, layout.row_gdp, columns),
            value_added_components_source=_vector_by_columns(
                ws_domestic, layout.row_value_added_components, columns
            ),
            compensation=_vector_by_columns(ws_domestic, layout.row_compensation, columns),
            taxes_production_imports=_vector_by_columns(
                ws_domestic, layout.row_taxes_production_imports, columns
            ),
            subsidies_production=_vector_by_columns(
                ws_domestic, layout.row_subsidies_production, columns
            ),
            operating_surplus=_vector_by_columns(
                ws_domestic, layout.row_operating_surplus, columns
            ),
            mixed_income=_vector_by_columns(ws_domestic, layout.row_mixed_income, columns),
            jobs=_vector_by_columns(ws_domestic, layout.row_jobs, columns),
        )
    finally:
        workbook.close()

